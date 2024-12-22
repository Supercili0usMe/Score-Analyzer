import unittest
import os
import openpyxl as xl
from datetime import datetime
from src.data_processing import *
from unittest.mock import patch, MagicMock
from unittest.mock import Mock

class TestDataProcessing(unittest.TestCase):
    def setUp(self):
        self.base_folder = "data"
        self.file_name = "Отметки_1.xlsx"
        self.project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        self.folder_root = os.path.join(self.project_root, self.base_folder)
        self.valid_file_folder = os.path.join(self.folder_root, self.file_name)
        self.invalid_file_folder = os.path.join(self.folder_root, "нету.xlsx")

    @patch("os.path.exists")
    def test_get_file_path_valid(self, mock_exists):
        '''Тестируем успешное получение пути к файлу.'''
        mock_exists.return_value = True
        result = get_file_path(self.file_name, self.base_folder)
        self.assertEqual(result, self.valid_file_folder)
    
    @patch("os.path.exists")
    def test_get_file_path_file_not_found(self, mock_exists):
        '''Тестируем поведение при отсутствии файла.'''
        mock_exists.return_value = False
        with self.assertRaises(FileNotFoundError):
            get_file_path(self.file_name, self.base_folder)

    @patch("openpyxl.load_workbook")
    def test_read_excel_valid(self, mock_read_excel):
        '''Тестируем успешное чтение файла.'''
        mock_read_excel.return_value = xl.Workbook()
        result = read_excel(self.valid_file_folder)
        self.assertIsInstance(result, xl.workbook.workbook.Workbook)

    @patch("openpyxl.load_workbook")
    def test_read_excel_error(self, mock_read_excel):
        '''Тестируем обработку ошибок при чтении файла.'''
        mock_read_excel.side_effect = Exception("Файл поврежден")
        with self.assertRaises(ValueError):
            read_excel(self.invalid_file_folder)
    
class TestExtractInfo(unittest.TestCase):
    def setUp(self):
        self.workbook = xl.Workbook()
        self.sheet = self.workbook.active

    def test_basic_case(self):
        '''Тестируем корректное извлечение информации.'''
        self.sheet["A1"] = "Организация:"
        self.sheet["A2"] = "Хогвардс"
        self.sheet["A3"] = "Обучающийся:"
        self.sheet["A4"] = "Гарри Владимирович Поттер"
        self.sheet["A5"] = "Класс:"
        self.sheet["A6"] = "Выпускной"
        self.sheet["A7"] = "Период:"
        self.sheet["A8"] = "Расцвет римской империи"

        expected = {"Организация": "Хогвардс", "Обучающийся": "Гарри Владимирович Поттер",
                    "Класс": "Выпускной", "Период": "Расцвет римской империи"}
        
        result = extract_info(self.sheet)
        self.assertEqual(result, expected)
    
    def test_empty_cells(self):
        '''Тестируем случай, когда ячейки пусты.'''
        self.sheet["A1"] = "Организация:"
        self.sheet["A2"] = "Хогвардс"
        self.sheet["A3"] = "Обучающийся:"
        self.sheet["A4"] = "Гарри Владимирович Поттер"
        self.sheet["A5"] = "Класс:"
        self.sheet["A6"] = None
        self.sheet["A7"] = None
        self.sheet["A8"] = None

        expected = {"Организация": "Хогвардс", "Обучающийся": "Гарри Владимирович Поттер",
                    "Класс": None}
        result = extract_info(self.sheet)
        self.assertEqual(result, expected)

    def test_invalid_cells(self):
        '''Тестируем случай, когда ключи или значения отсутствуют'''
        self.sheet["A1"] = None
        self.sheet["A2"] = "Хогвардс"
        self.sheet["A3"] = "Обучающийся:"
        self.sheet["A4"] = "Гарри Владимирович Поттер"
        self.sheet["A5"] = None
        self.sheet["A6"] = None
        self.sheet["A7"] = None
        self.sheet["A8"] = None

        expected = {"Обучающийся": "Гарри Владимирович Поттер"}
        result = extract_info(self.sheet)
        self.assertEqual(result, expected)

class TestExtractSubjects(unittest.TestCase):
    def setUp(self):
        self.workbook = xl.Workbook()
        self.sheet = self.workbook.active

    def test_basic_case(self):
        '''Тестируем корректное извлечение информации.'''
        self.sheet["A1"] = 'Изобразительное искусство'
        self.sheet["A2"] = "Иностранный язык"
        self.sheet["A3"] = "Литературное чтение"
        self.sheet["A4"] = "Математика"
        self.sheet["A5"] = "Музыка"
        self.sheet["A6"] = "Окружающий мир"
        self.sheet["A7"] = None

        expected = {1: 'Изобразительное искусство',
                    2: "Иностранный язык",
                    3: "Литературное чтение",
                    4: "Математика",
                    5: "Музыка",
                    6: "Окружающий мир"}
        result = extract_subjects(self.sheet, 1)
        self.assertEqual(result, expected)

    def test_empty_cells(self):
        '''Тестируем случай пустых данных'''
        self.sheet["A1"] = None
        self.sheet["A2"] = None
        self.sheet["A3"] = None
        self.sheet["A4"] = None
        self.sheet["A5"] = None
        self.sheet["A6"] = None
        self.sheet["A7"] = None

        expected = {}
        result = extract_subjects(self.sheet, 0)
        self.assertEqual(result, expected)

class TestExtractMarks(unittest.TestCase):
    def setUp(self):
        self.worksheet = Mock()
        self.subjects = {1: "Math", 2: "Physics"}

    def test_basic_marks_extraction(self):
        self.worksheet.iter_cols.return_value = [
            (datetime(2023, 1, 1), "5", "4"),
            (datetime(2023, 1, 2), "4", "3")
        ]
        
        expected = {
            "Math": [
                {"Дата": datetime(2023, 1, 1), "Отметка": "5"},
                {"Дата": datetime(2023, 1, 2), "Отметка": "4"}
            ],
            "Physics": [
                {"Дата": datetime(2023, 1, 1), "Отметка": "4"},
                {"Дата": datetime(2023, 1, 2), "Отметка": "3"}
            ]
        }
        
        result = extract_marks(self.worksheet, self.subjects)
        self.assertEqual(result, expected)

    def test_multiple_marks_same_day(self):
        self.worksheet.iter_cols.return_value = [
            (datetime(2023, 1, 1), "54", "43")
        ]
        
        expected = {
            "Math": [
                {"Дата": datetime(2023, 1, 1), "Отметка": "5"},
                {"Дата": datetime(2023, 1, 1), "Отметка": "4"}
            ],
            "Physics": [
                {"Дата": datetime(2023, 1, 1), "Отметка": "4"},
                {"Дата": datetime(2023, 1, 1), "Отметка": "3"}
            ]
        }
        
        result = extract_marks(self.worksheet, self.subjects)
        self.assertEqual(result, expected)

    def test_non_numeric_marks(self):
        self.worksheet.iter_cols.return_value = [
            (datetime(2023, 1, 1), "н5", "4б")
        ]
        
        expected = {
            "Math": [
                {"Дата": datetime(2023, 1, 1), "Отметка": "5"}
            ],
            "Physics": [
                {"Дата": datetime(2023, 1, 1), "Отметка": "4"}
            ]
        }
        
        result = extract_marks(self.worksheet, self.subjects)
        self.assertEqual(result, expected)

    def test_итог_column(self):
        self.worksheet.iter_cols.return_value = [
            (datetime(2023, 1, 1), "5", "4"),
            ("Итог:", "5", "4")
        ]
        
        expected = {
            "Math": [
                {"Дата": datetime(2023, 1, 1), "Отметка": "5"}
            ],
            "Physics": [
                {"Дата": datetime(2023, 1, 1), "Отметка": "4"}
            ]
        }
        
        result = extract_marks(self.worksheet, self.subjects)
        self.assertEqual(result, expected)

    def test_empty_marks(self):
        self.worksheet.iter_cols.return_value = [
            (datetime(2023, 1, 1), "", None)
        ]
        
        expected = {
            "Math": [],
            "Physics": []
        }
        
        result = extract_marks(self.worksheet, self.subjects)
        self.assertEqual(result, expected)

class TestRefactorMarks(unittest.TestCase):
    def setUp(self):
        self.test_marks = {
            "Math": [
                {"Дата": datetime(2023, 9, 1), "Отметка": "5"},
                {"Дата": datetime(2023, 9, 1), "Отметка": "4"},
                {"Дата": datetime(2023, 9, 2), "Отметка": "3"}
            ],
            "Physics": [
                {"Дата": datetime(2023, 9, 1), "Отметка": "4"}
            ],
            "Chemistry": []
        }

    def test_valid_subject_multiple_marks(self):
        dates, grades = refactor_marks(self.test_marks, "Math")
        self.assertEqual(len(dates), 3)
        self.assertEqual(len(grades), 3)
        self.assertEqual(grades, ["5", "4", "3"])
        self.assertEqual(dates, [
            datetime(2023, 9, 1),
            datetime(2023, 9, 1),
            datetime(2023, 9, 2)
        ])

    def test_valid_subject_single_mark(self):
        dates, grades = refactor_marks(self.test_marks, "Physics")
        self.assertEqual(len(dates), 1)
        self.assertEqual(len(grades), 1)
        self.assertEqual(grades, ["4"])
        self.assertEqual(dates, [datetime(2023, 9, 1)])

    def test_empty_marks_list(self):
        dates, grades = refactor_marks(self.test_marks, "Chemistry")
        self.assertEqual(len(dates), 0)
        self.assertEqual(len(grades), 0)
        self.assertEqual(grades, [])
        self.assertEqual(dates, [])

    def test_nonexistent_subject(self):
        with self.assertRaises(KeyError):
            refactor_marks(self.test_marks, "Biology")

    def test_empty_marks_dict(self):
        empty_marks = {}
        with self.assertRaises(KeyError):
            refactor_marks(empty_marks, "Math")

if __name__ == "__main__":
    unittest.main()