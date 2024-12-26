import os
from openpyxl import load_workbook, Workbook
import openpyxl

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
folder_root = os.path.join(project_root, 'data')
coeffs = {
    "Административная контрольная работа": 1.5,     "Аудирование": 1.4,
    "Ведение тетради": 1,                           "Дистанционное занятие": 1,
    "Дифференцированный зачет": 1,                  "Домашнее задание": 1,
    "Домашнее сочинение": 1.4,                      "Зачёт": 1.5,
    "Изложение": 1.4,                               "Инструктаж": 1,
    "Квалификационное испытание": 1,                "Классное сочинение": 1.5,
    "Классное сочинение": 1.5,                      "Контрольная практическая работа": 1.5,
    "Контрольная работа": 1.5,                      "Контрольный диктант": 1.5,
    "Курсовая работа": 1,                           "Лабораторная работа": 1.3,
    "Практическая работа": 1.3,                     "Проверочная работа": 1.3,
    "Проект": 1,                                    "Работа на занятии": 1,
    "Работа на уроке": 1,                           "Самостоятельная работа": 1.2,
    "Словарный диктант": 1.4,                       "Сочинение": 1,
    "Срезовая работа": 1.3,                         "Тест": 1,
    "Чтение наизусть": 1,                           "Электронное обучение": 1,
}

def get_file_path(file_name: str, base_folder: str) -> str:
    '''Возвращает полный путь к файлу, если он существует в указанной папке'''
    file_path = os.path.join(folder_root, file_name)
    if not os.path.exists(file_path):
        raise FileNotFoundError(f'Файл "{file_name}" не найден в папке "{base_folder}"')
    return file_path

def read_excel(file_path: str) -> Workbook:
    '''Читает Excel-файл и возвращает его содержимое как `openpyxl.Workbook`, не Worksheet!'''
    try:
        data = load_workbook(file_path)
        return data
    except Exception as e:
        raise ValueError(f"Ошибка при чтении файла: {e}")
    
def extract_info(worksheet, start_row=1, step=2, max_rows=7) -> dict:
    '''
    Извлекаем информацию об ученике из открытого листа Excel
    
    Аргументы:
        `worksheet`: Лист Excel
        `start_row`: Номер начальной строки, с которой нужно собирать инфу (default: 1)
        `step`: Количество строк, пропускаемое для корректного сбора (default: 2)
        `max_rows`: Последняя строка из списка (default: 7)
    
    Возвращает:
        Словарь вида `{type: param}`, где `type` - основные характеристики,
        `param` - данные по этим характеристикам
    '''
    info = {}
    for row in range(start_row, max_rows + start_row, step):
        key_cell = worksheet[f"A{row}"]
        value_cell = worksheet[f"A{row+1}"]

        # Извлекаем данные из ячеек
        key = key_cell.value[:-1] if key_cell.value and isinstance(key_cell.value, str) else None
        value = value_cell.value if value_cell.value is not None else None

        if key:
            info[key] = value
    return info

def extract_subjects(worksheet, start_row=11, column_index=0) -> dict:
    """
    Выделяем список предметов из открытого листа Excel
    
    Аргументы:
        `worksheet`: Лист Excel
        `start_row`: Строка, с которой начинается изъятие (default: 11)
        `column_index`: Столбец, в котором находятся предметы (default: 0)
    
    Возвращает:
        Словарь вида `{num: subj}`, где `num` - номер предмета, начиная от 1,
        `subj` - название предмета
    """
    subjects = {}
    for idx, row in enumerate(list(worksheet.rows)[start_row-1:]):
        if row[column_index].value is None:
            break
        subjects[idx + 1] = row[column_index].value
    return subjects

def extract_marks(worksheet, subjects: dict, start_row=10, start_column=2) -> dict:
    """
    Получаем список оценок для каждого предмета из листа Excel.
    
    Аргументы:
        `worksheet`: Лист Excel
        `subjects`: Словарь предметов, получаемый в результате работы функции `extract_subjects`
        `start_row`: Строка, с которой начинается изъятие (default: 10)
        `start_column`: Столбец, с которого начинает осмотр оценок (default: 2)
    
    Возвращает:
        Словарь вида `{subj: [{"дата": date, "отметка": mark, "Тип работы": type, "Коэффициент": coeff}]}`, где subj - название предмета,
        date - дата получения отметки, mark - сама отметка, type - тип полученной отметки, coeff - коэффициент отметки.
    
    Важно:
        Если в один день по одному предмету больше 1 отметки, то записывается каждая из отметок
    """
    # Создаём словарь из предметов
    marks = {subj: [] for subj in subjects.values()}

    # Указываем параметры
    start_row = 10
    start_column = 2

    for value in worksheet.iter_cols(min_row=start_row, max_row=start_row+len(subjects), min_col=start_column):
        # Забираем дату из первой строки
        date = value[0].value
        if date == "Итог:": break

        # Проходимся по остальным строкам
        for subj_id, cell in enumerate(value[1:]):
            # Если ячейка пустая - пропускаем
            if cell is None or not cell.value or not cell.comment:
                continue
            # Проходимся по всем отметкам и комментариям одновременно
            for mark, comment in zip(cell.value, cell.comment.text.strip().split(";")):
                # Если символ не числа - пропускаем
                if not mark.isdigit():
                    continue
                # Выделяем комментарии
                comment = comment.strip()
                # Получаем тип отметки и саму отметку
                if comment:
                    _, work_type, _ = comment.split(" - ")
                    mark_data = {
                        "Дата": date, 
                        "Отметка": int(mark),
                        "Тип работы": work_type,
                        "Коэффициент": coeffs[work_type]
                    }
                marks[subjects[subj_id+1]].append(mark_data)
    return marks

def refactor_marks(marks: dict, subject: str) -> tuple[list, list]:
    """
    Выделяем из словаря отметок массив дат и массив оценок
    
    Аргументы:
        `marks`: Словарь отметок, получаемый в результате выполнения функции `refactor_marks`
        `subject`: Строка - название получаемого предмета
    
    Возвращает:
        Три массива: [dates], [marks] и [coeffs], отвечающие за даты, отметки и коэффициенты соответственно 
    
    Важно:
        Если в один день по одному предмету больше 1 отметки, то записывается каждая из отметок и каждая дата
    """
    dates, grades, coeffs = [], [], []
    for info in marks[subject]:
        dates.append(info["Дата"])
        grades.append(info["Отметка"])
        coeffs.append(info["коэффициент"])
    return dates, grades, coeffs

def main():
    # Проверяем существование папки
    if not os.path.exists(folder_root):
        print(f"Папка 'data' не найдена по пути: {folder_root}")
        return

    # Запрос имени файла у пользователя
    file_name = input("Введите имя файла (например, example.xlsx): ").strip()

    try:
        file_path = get_file_path(file_name, folder_root)
        data = read_excel(file_path)
        print("Данные успешно загружены!")
        return data
    except (FileNotFoundError, ValueError) as e:
        print(e)

if __name__ == "__main__":
    data = main()

